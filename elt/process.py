import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# process mtr

mtr_df = pd.read_excel("./files/Merchant_Tax_Report_(MTR)_Sheet_-_Hiring.xlsx")

mtr_df = mtr_df[mtr_df["Transaction Type"] != "Cancel"]

mtr_df["Transaction Type"] = mtr_df["Transaction Type"].replace("Refund", "Return")

mtr_df["Transaction Type"] = mtr_df["Transaction Type"].replace(
    "FreeReplacement", "Return"
)

mtr_df.to_csv("./files/Transformed_MTR.csv", index=False)

# process payment report

payment_df = pd.read_csv("./files/Payment_Report_Sheet_-_Hiring_-_Sheet1.csv")

payment_df = payment_df[payment_df["type"] != "Transfer"]

payment_df.rename(columns={"type": "Payment Type"}, inplace=True)

payment_df["Payment Type"] = payment_df["Payment Type"].replace(
    {
        "Ajdustment": "Order",
        "FBA Inventory Fee": "Order",
        "Fulfilment Fee Refund": "Order",
        "Service Fee": "Order",
        "Refund": "Return",
    }
)

payment_df["Transaction Type"] = "Payment"

payment_df.to_csv("./files/Transformed_Payment_Report.csv", index=False)

# merge

wb = Workbook()
ws = wb.active
ws.title = "Merged Sheet"

orange_fill = PatternFill(start_color="ff7b00", end_color="ff7b00", fill_type="solid")
blue_fill = PatternFill(start_color="002a60", end_color="002a60", fill_type="solid")
white_font = Font(color="FFFFFF")

columns = [
    "Order Id",
    "Transaction Type",
    "Payment Type",
    "Invoice Amount",
    "Net Amount",
    "P_Description",
    "Order Date",
    "Payment Date",
]
cols = ["mtr", "mtr", "pay", "mtr", "pay", "pay", "mtr", "pay"]

ws.append(columns)

for i, cell in enumerate(ws[ws.max_row]):
    if cols[i] == "mtr":
        cell.fill = orange_fill
    elif cols[i] == "pay":
        cell.fill = blue_fill
        cell.font = white_font


def map_mtr_row(row):
    return {
        "Order Id": row.get("Order Id", ""),
        "Transaction Type": row.get("Transaction Type", ""),
        "Payment Type": "",
        "Invoice Amount": row.get("Invoice Amount", ""),
        "Net Amount": "",
        "P_Description": "",
        "Order Date": row.get("Order Date", ""),
        "Payment Date": "",
    }


def map_payment_row(row):
    return {
        "Order Id": row.get("order id", ""),
        "Transaction Type": row.get("Transaction Type", ""),
        "Payment Type": row.get("Payment Type", ""),
        "Invoice Amount": "",
        "Net Amount": row.get("total", ""),
        "P_Description": row.get("description", ""),
        "Order Date": "",
        "Payment Date": row.get("date/time", ""),
    }


for _, row in mtr_df.iterrows():
    mapped_row = map_mtr_row(row)
    ws.append(
        [
            (
                mapped_row[col].strip()
                if isinstance(mapped_row[col], str)
                else mapped_row[col]
            )
            for col in columns
        ]
    )
    for cell in ws[ws.max_row]:
        cell.fill = orange_fill

for _, row in payment_df.iterrows():
    mapped_row = map_payment_row(row)
    ws.append(
        [
            (
                mapped_row[col].strip()
                if isinstance(mapped_row[col], str)
                else mapped_row[col]
            )
            for col in columns
        ]
    )
    for cell in ws[ws.max_row]:
        cell.fill = blue_fill
        cell.font = white_font

wb.save("./files/Merged_Report5.xlsx")


print("Merged Excel file has been created successfully.")
